import pandas as pd
import numpy as np  
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
# Read the all Employee file
df1 = pd.read_csv('Employee_20250901094051.csv')
#Read the First and Last file
df2 = pd.read_csv('30_08_2025\First & Last_20250901090917_export.csv')
#Read the Leave Report file
df3 = pd.read_csv('30_08_2025\Leave Report_20250901090253_export.csv')

# #Ensure the   'EmployeeID' column is treated  consistently
df1.columns = df1.columns.str.strip() 
df2.columns = df2.columns.str.strip()
df3.columns = df3.columns.str.strip()
print("df1 columns:", df1.columns.tolist())
print("df2 columns:", df2.columns.tolist())
print("df3 columns:", df3.columns.tolist())
df1['Employee ID'] = df1['Employee ID'].astype(str)
df2['Employee ID'] = df2['Employee ID'].astype(str)
df3['Employee ID'] = df3['Employee ID'].astype(str)

#Find records in df1 that are not in df2
missing_records1 = df1[~df1['Employee ID'].isin(df2['Employee ID'])]
missing_records=missing_records1[~missing_records1['Employee ID'].isin(df3['Employee ID'])].sort_values(by='Department')
# Create an empty list to store new rows 
header_rows = []
grouped_with_counts = []

# Create the empty row as the first row
head_row = {col: '' for col in missing_records.columns}
if 'COUNT' not in head_row:
  head_row['COUNT'] = ''
prev_day = (datetime.now() - timedelta(days=1)).strftime('%A %d/%m/%Y')
head_row['Department'] = f"ABSENTEES REPORT {prev_day}"
grouped_with_counts.append(pd.DataFrame([head_row]))  # Add empty row as the first row

# Group by Department
for dept, group in missing_records.groupby('Department'):
  
  #create a empty row that is centre merged
    empty_row = {col: '' for col in missing_records.columns}
    grouped_with_counts.append(pd.DataFrame([empty_row])) 
    # Create a count row
    count_row = {col: '' for col in missing_records.columns}
    count_row['Department'] = f"Department: {dept}"
    count_row['COUNT'] = f"Total Count: {len(group)}"
  
    # Add count row before the group, with yellow background
    count_row_df = pd.DataFrame([count_row])
    # No need to style with pandas here; we'll use openpyxl after saving the Excel file.
    grouped_with_counts.append(count_row_df)
    grouped_with_counts.append(group)  # Add all rows of the group

# Combine all into one DataFrame
final_df = pd.concat(grouped_with_counts, ignore_index=True)
# Save the missing records to a new excel sheet file
final_df.to_excel('missing_records07sat.xlsx', index=False)

excel_path = 'missing_records07sat.xlsx'
wb = load_workbook(excel_path)
ws = wb.active
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    # Check if the cell in the 'Department' column starts with "Department:"
    # Find the index of the 'Department' column
    dept_col_idx = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == 'Department':
            dept_col_idx = idx
            break
    if dept_col_idx is not None and str(row[dept_col_idx].value).startswith("Department:"):
        for cell in row:
            cell.fill = yellow_fill

wb.save(excel_path)

#display the missing records
print("Missing records saved to 'missing_records.xlsx'.")