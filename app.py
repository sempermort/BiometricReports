from flask import Flask, render_template, request, send_file, redirect, url_for, flash
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import os
import webbrowser
import threading

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
BASE_DIR = app.root_path 
# Ensure upload and output folders exist
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER


for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# Route for the home page
@app.route('/')
def index():
    return render_template('index.html')



# Route for the summary report results page
@app.route('/summaryresults')
def summaryresults():
    return render_template('summaryresults.html')
# Route for the missing people report results page
@app.route('/missingresults')
def missingresults():
    return render_template('missingresults.html')
@app.route('/presentleavelateresults')
def presentleavelateresults():
    return render_template('presentleavelatein.html')


# Route for the summary report tab
@app.route('/mis')
def summary_report():
    # Check if summary report exists
    summary_file_path = os.path.join(app.config['OUTPUT_FOLDER'], 'summary_report.xlsx')
    if os.path.exists(summary_file_path):
        return render_template('summary.html', summary_file=summary_file_path)
    else:
        return render_template('summary.html')
    

def process_department_file(df, output_path):
    """
    Processes a DataFrame, writes it to Excel, fills department header rows with yellow,
    and adds a COUNT column with the number of rows for each department.
    """
    df.to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active


    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
     for cell in row:
        if cell.value and str(cell.value).startswith("Department:"):
            # Fill the row yellow
            for c in row:
                c.fill = yellow_fill
            count = 0
            current_row_idx = cell.row + 1
            while current_row_idx <= ws.max_row:
                next_row = ws[current_row_idx]
                found_next_header = False
                for next_cell in next_row:
                    if next_cell.value and str(next_cell.value).startswith("Department:"):
                        found_next_header = True
                        break
                if found_next_header:
                    break
                # Count if column 1 (index 0) is a number
                first_cell = next_row[0]
                if isinstance(first_cell.value, (int, float)) and first_cell.value != '':
                    count += 1
                current_row_idx += 1
            # Write the count in the last column of the header row
            ws.cell(row=cell.row, column=ws.max_column, value=f'Count: {count}')
            break  # Only process once per header row
    wb.save(output_path)

@app.route('/process_present_leave_late', methods=['POST'])
def process_present_leave_late():
    try:
        present_file_df = pd.read_excel(request.files.get('present_file_dw'))
        leave_file_df = pd.read_excel(request.files.get('Leave_file_dw'))
        latein_file_df = pd.read_excel(request.files.get('late_in_file_dw'))

        # Check if files were uploaded
        if not all([present_file_df is not None, leave_file_df is not None, latein_file_df is not None]):
            flash('Please upload all required files')
            return redirect(url_for('index'))

        # Use the helper function for each file
        process_department_file(present_file_df, os.path.join(app.config['OUTPUT_FOLDER'], 'present_records_report.xlsx'))
        process_department_file(leave_file_df, os.path.join(app.config['OUTPUT_FOLDER'], 'leave_records_report.xlsx'))
        process_department_file(latein_file_df, os.path.join(app.config['OUTPUT_FOLDER'], 'latein_records_report.xlsx'))

        return redirect(url_for('presentleavelateresults'))
    except Exception as e:
        flash(f'Error processing files: {str(e)}')
        return redirect(url_for('index'))


# Route for handling file uploads and processing
@app.route('/process_missing', methods=['POST'])
def process_files():
    try:
        # Get uploaded files
        employee_file = request.files.get('employee_file')
        first_last_file = request.files.get('first_last_file')
        leave_report_file = request.files.get('leave_report_file')
        
        # Check if files were uploaded
        if not all([employee_file, first_last_file, leave_report_file]):
            flash('Please upload all required files')
            return redirect(url_for('index'))
        
        #Read the all Employee file
        df1 = pd.read_csv(employee_file)
        #Read the First and Last file
        df2 = pd.read_csv(first_last_file)
        #Read the Leave Report file
        df3 = pd.read_csv(leave_report_file)

        # --- Step 2: Find the 'Date' column ---
        if "Date" not in df2.columns:
            raise ValueError("❌ No 'Date' column found in CSV!")

        # Get the index of the Date column
        date_col_idx = df2.columns.get_loc("Date")

        # --- Step 3: Get the value below the column header (second row) ---
        raw_date = df2.iloc[0, date_col_idx]  # First data row under the 'Date' column

        date_obj = datetime.strptime(str(raw_date), "%Y-%m-%d")


        # #Ensure the   'EmployeeID' column is treated  consistently
        df1.columns = df1.columns.str.strip()
        df2.columns = df2.columns.str.strip()
        df3.columns = df3.columns.str.strip()
        df1['Employee ID'] = df1['Employee ID'].astype(str)
        df2['Employee ID'] = df2['Employee ID'].astype(str)
        df3['Employee ID'] = df3['Employee ID'].astype(str)

        #Find records in df1 that are not in df2
        missing_records1 = df1[~df1['Employee ID'].isin(df2['Employee ID'])]
        missing_records=missing_records1[~missing_records1['Employee ID'].isin(df3['Employee ID'])].sort_values(by='Department')

        # Create an empty list to store new rows
        header_rows = []
        grouped_with_counts = []

        # Create the empty row as the first row
        head_row = {col: '' for col in missing_records.columns}
        if 'COUNT' not in head_row:
            head_row['COUNT'] = ''
      
        head_row['Department'] = f"ABSENTEES REPORT {date_obj.strftime('%A %d/%m/%Y')}"
        grouped_with_counts.append(pd.DataFrame([head_row]))  # Add empty row as the first row

        # Group by Department
        for dept, group in missing_records.groupby('Department'):
            
            #create a empty row that is centre merged
            empty_row = {col: '' for col in missing_records.columns}
            grouped_with_counts.append(pd.DataFrame([empty_row]))
            # Create a count row
            count_row = {col: '' for col in missing_records.columns}
            count_row['Department'] = f"Department: {dept}"
            count_row['COUNT'] = f"Total Count: {len(group)}"
            
            # Add count row before the group, with yellow background
            count_row_df = pd.DataFrame([count_row])

            # No need to style with pandas here; we'll use openpyxl after saving the Excel file.
            grouped_with_counts.append(count_row_df)
            grouped_with_counts.append(group)  # Add all rows of the group

        # Combine all into one DataFrame
        final_df = pd.concat(grouped_with_counts, ignore_index=True)
        
        # Save the final DataFrame to an Excel file
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], 'missing_records_report.xlsx')
        
        final_df.to_excel(output_path, index=False)

           
        wb = load_workbook(output_path)
        ws = wb.active

        # Swap the first and second rows
        first_row = [cell.value for cell in ws[1]]
        second_row = [cell.value for cell in ws[2]]
        #first row with date
        end_col = len(ws[1])

        value_found = None

        # Loop through the second row
        for cell in ws[2]:  # ws[2] is the second row
            if cell.value is not None and str(cell.value).strip() != "":
                value_found = cell.value
                break
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)

        # Set the value of the merged cell (use a single string)
        ws.cell(row=1, column=1, value=value_found)  # joins list nto string
            

        for col_idx, value in enumerate(first_row, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        #fill the color to the departiment header
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        dept_col_idx = None
        for idx, cell in enumerate(ws[2]):
                        if cell.value == 'Department':
                            dept_col_idx = idx
                            break
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            # Check if the cell in the 'Department' column starts with "Department:"
            # Find the index of the 'Department' column
           
           
            if dept_col_idx is not None and str(row[dept_col_idx].value).startswith("Department:"):
                for cell in row:
                    cell.fill = yellow_fill

        wb.save(output_path)

        # Return the processed files
        return  redirect(url_for('missingresults'))
    except Exception as e:
        flash(f'Error processing files: {str(e)}')
        return redirect(url_for('index'))


@app.route('/process_summary_report', methods=['POST'])
def process_summary_report():
    try:
        # Get uploaded files
        employee_file_excel = request.files.get('employee_file_excel')
        first_last_file_excel = request.files.get('first_last_file_excel')
        leave_report_file_excel = request.files.get('leave_report_file_excel')
        
        if not all([employee_file_excel,  first_last_file_excel, leave_report_file_excel]):
            flash('Please upload all required files')
            return redirect(url_for('index'))
      
        # Step 1: Read source Excel First and last file
        df = pd.read_excel(first_last_file_excel, skiprows=2)
        grouped = df.groupby('Department').size().reset_index(name='Count')

        # Step 2: Read source Excel leave file
        df2 = pd.read_excel(leave_report_file_excel, skiprows=2)
        grouped2 = df2.groupby('Department').size().reset_index(name='Count')

        # Step 3: Read source Excel All Employee file
        df3 = pd.read_excel(employee_file_excel, skiprows=1)
        grouped3 = df3.groupby('Department').size().reset_index(name='Count')

        # Load the first and last workbook to get the date
        flwb = load_workbook(first_last_file_excel)

        # Get the active  first and last worksheet (assuming the date is in the active sheet) 
        first_last_ws = flwb.active  

        # Get the date 
        date = first_last_ws["A2"].value  # You can change A2 to any column you want

        date_str = date.replace("Date: ", "")
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
       
        # Step 4: Load summary Excel file
        summary_file = os.path.join(BASE_DIR, "summary", "sumaryreport.xlsx")
        wb = load_workbook(summary_file)
        ws = wb.active

        # Set this value into cell E2 of the target worksheet
        ws["E2"] = date_obj

        # Step 5: Map department keywords to department names in grouped
        department_map = {
            'PERMANENT AGRICULTURE': 'AGRICULTURE',
            'CASUAL  AGRIC': 'CASUAL  AGRIC',
            'PERMANENT IRRIGATION': 'IRRIGATION',
            'CASUAL IRRIGATION': 'CASUAL IRRIGATION',
            'PERMANENT ADMINISTRATION': 'ADMINISTRATION',
            'CASUAL ADMINISTRATION': 'CASUAL ADMINISTRATION',
            'PERMANENT CIVIL': 'CIVIL',
            'CASUAL CIVIL': 'CASUAL CIVIL',
            'WORKSHOP PERMANENT': 'WORKSHOP',
            'CASUAL WORKSHOP': 'CASUAL WORKSHOP',
            'LAND DEVELOPMENT PERMANENT': 'LAND DEVELOPMENT',
            'CASUAL LAND DEVELOPMENT': 'CASUAL LAND DEVELOPMENT',
            'SURVEYOR-PERMANENT': 'SURVEYOR',
            'CASUAL -SURVEYOR': 'CASUAL SURVEYOR',
            'STORE-PERMANENT': 'STORE',
            'STORE-CASUAL': 'CASUAL STORE',
            'ELECTRICAL-PERMANENT': 'ELECTRICAL',
            'ELECTRICAL-CASUAL': 'CASUAL ELECTRICAL',
            'SECURITY-PERMANENT': 'SECURITY',
            'SECURITY-CASUAL': 'CASUAL SECURITY',
            'SHEQ-PERMANENT': 'SHEQ',
            'SHEQ-CASUAL': 'CASUAL SHEQ',
            'IT-PERMANENT': 'IT',
            'IT-CASUAL': 'CASUAL IT',
            'HR-PERMANENT': 'HUMAN RESOURCES',
            'HR-CASUAL': 'CASUAL HUMAN RESOURCES',
            'MANAGERS AND EXPERTIES': 'MANAGERS & EXPERTS',
            'AIR SECURITY': 'AIR SECURITY',
            'INTERN/GRADUATES': 'INTERN/GRADUATES',
            'FACTORY-PERMANENT': 'FACTORY',
            'FACTORY-CASUAL': 'CASUAL FACTORY',
            'TRANSPORT-PERMANENT': 'TRANSPORT',
            'TRANSPORT-CASUAL': 'CASUAL TRANSPORT',
            'PROCUREMENT-PERMANENT': 'PROCUREMENT',
            'PROCUREMENT-CASUAL': 'CASUAL PROCUREMENT'
        }

        # Step 6: Iterate through rows in ws and fill columns 3, 5, and 7 based on mapping
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            cell = row[0]
            if cell.value:
                for key, dept in department_map.items():
                    if key in str(cell.value):
                        # Fill column 3 from source ALL employee file
                        dept_count3 = grouped3.loc[grouped3['Department'] == dept, 'Count']
                        if not dept_count3.empty:
                            ws.cell(row=cell.row, column=3, value=int(dept_count3.iloc[0]))
                        else:
                            ws.cell(row=cell.row, column=3, value="")
                        # Fill column 5 from source first and last file
                        dept_count = grouped.loc[grouped['Department'] == dept, 'Count']
                        if not dept_count.empty:
                            ws.cell(row=cell.row, column=5, value=int(dept_count.iloc[0]))
                        else:
                            ws.cell(row=cell.row, column=5, value="")
                        # Fill column 7 from source leave file
                        dept_count2 = grouped2.loc[grouped2['Department'] == dept, 'Count']
                        if not dept_count2.empty:
                            ws.cell(row=cell.row, column=7, value=int(dept_count2.iloc[0]))
                        else:
                            ws.cell(row=cell.row, column=7, value="")
                        break  # Stop after first match        
        
        # Define output path for the processed file
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], 'summary_report.xlsx')
        wb.save(output_path)
        # Copy the processed file to output folder if needed
        # For now, we'll just return the summary file path
        return  redirect(url_for('summaryresults'))
    except Exception as e:
        flash(f'Error processing summary report: {str(e)}')
        return redirect(url_for('index'))

@app.route('/download/<path:filename>')
def download_file(filename):
    file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        flash('File not found')
        return redirect(url_for('index'))

if __name__ == "__main__":

    import webbrowser
    import threading

    # Open browser automatically
    def open_browser():
        webbrowser.open("http://127.0.0.1:5000/")

    threading.Timer(1.0, open_browser).start()
    app.run(debug=False)
