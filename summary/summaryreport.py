import pandas as pd
from openpyxl import load_workbook

# Step 1: Read source Excel First and last file
source_file = 'First & Last_20250822090342_export.xlsx'
df = pd.read_excel(source_file, skiprows=2)
grouped = df.groupby('Department').size().reset_index(name='Count')

# Step 2: Read source Excel leave file
source_file2 = 'Leave Report_20250822084651_export.xlsx'
df2 = pd.read_excel(source_file2, skiprows=2)
grouped2 = df2.groupby('Department').size().reset_index(name='Count')

# Step 3: Read source Excel All Employee file
source_file3 = 'Employee_20250822091528.xlsx'  # Change to your actual third file name
df3 = pd.read_excel(source_file3, skiprows=1)
grouped3 = df3.groupby('Department').size().reset_index(name='Count')

# Step 4: Load summary Excel file
summary_file = 'sumaryreport.xlsx'
wb = load_workbook(summary_file)
ws = wb.active

# Step 5: Map department keywords to department names in grouped
department_map = {
    'PERMANENT AGRICULTURE': 'AGRICULTURE',
    'CASUAL  AGRIC': 'CASUAL  AGRIC',
    'PERMANENT IRRIGATION': 'IRRIGATION',
    'CASUAL IRRIGATION': 'CASUAL IRRIGATION',
    'PERMANENT ADMINISTRATION': 'ADMINISTRATION',
    'CASUAL ADMINISTRATION': 'CASUAL ADMINISTRATION',
    'PERMANENT CIVIL': 'CIVIL',
    'CASUAL CIVIL': 'CASUAL CIVIL',
    'WORKSHOP PERMANENT': 'WORKSHOP',
    'CASUAL WORKSHOP': 'CASUAL WORKSHOP',
    'LAND DEVELOPMENT PERMANENT': 'LAND DEVELOPMENT',
    'CASUAL LAND DEVELOPMENT': 'CASUAL LAND DEVELOPMENT',
    'SURVEYOR-PERMANENT': 'SURVEYOR',
    'CASUAL -SURVEYOR': 'CASUAL SURVEYOR',
    'STORE-PARMANENT': 'STORE',
    'STORE-CASUAL': 'CASUAL STORE',
    'ELECTRICAL-PARMANENT': 'ELECTRICAL',
    'ELECTRICAL-CASUAL': 'CASUAL ELECTRICAL',
    'SECURITY-PARMANENT': 'SECURITY',
    'SECURITY-CASUAL': 'CASUAL SECURITY',
    'SHEQ-PARMANENT': 'SHEQ',
    'SHEQ-CASUAL': 'CASUAL SHEQ',
    'IT-PARMANENT': 'IT',
    'IT-CASUAL': 'CASUAL IT',
    'HR-PARMANENT': 'HUMAN RESOURCES',
    'HR-CASUAL': 'CASUAL HUMAN RESOURCES',
    'MANAGERS AND EXPERTIES': 'MANAGERS & EXPERTS',
    'AIR SECURITY': 'AIR SECURITY',
    'INTERN/GRADUATES': 'INTERN/GRADUATES'
}

# Step 6: Iterate through rows in ws and fill columns 3, 5, and 7 based on mapping
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    cell = row[0]
    if cell.value:
        for key, dept in department_map.items():
            if key in str(cell.value):
                # Fill column 3 from source ALL employee file
                dept_count3 = grouped3.loc[grouped3['Department'] == dept, 'Count']
                if not dept_count3.empty:
                    ws.cell(row=cell.row, column=3, value=int(dept_count3.iloc[0]))
                else:
                    ws.cell(row=cell.row, column=3, value="")
                # Fill column 5 from source first and last file
                dept_count = grouped.loc[grouped['Department'] == dept, 'Count']
                if not dept_count.empty:
                    ws.cell(row=cell.row, column=5, value=int(dept_count.iloc[0]))
                else:
                    ws.cell(row=cell.row, column=5, value="")
                # Fill column 7 from source leave file
                dept_count2 = grouped2.loc[grouped2['Department'] == dept, 'Count']
                if not dept_count2.empty:
                    ws.cell(row=cell.row, column=7, value=int(dept_count2.iloc[0]))
                else:
                    ws.cell(row=cell.row, column=7, value="")
                break  # Stop after first match

wb.save(summary_file)
print("Summary report updated successfully.")